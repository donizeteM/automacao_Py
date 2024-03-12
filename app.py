import openpyxl
import pyperclip
import pyautogui
from time import sleep       # esse serve pra colocar uma pausa

workbook = openpyxl.load_workbook('produtos_ficticios.xlsx')
sheet_produtos = workbook['Produtos']
## iter_rows, passa linha por linha e min_row=2 entra na planilha e pega os dados da linha 2
for linha in sheet_produtos.iter_rows(min_row=2):
    nome_produto = linha[0].value
    pyperclip.copy(nome_produto)
    pyautogui.click(135,240,duration=1)     # clicou nas cordena
    pyautogui.hotkey('ctrl', 'v')            # fazendo um colar
    
    descricao = linha[1].value
    pyperclip.copy(descricao)
    pyautogui.click(138,341,duration=1) 
    pyautogui.hotkey('ctrl', 'v')
    
    categoria = linha[2].value
    pyperclip.copy(categoria)
    pyautogui.click(135,438,duration=1) 
    pyautogui.hotkey('ctrl', 'v')
    
    codigo_do__produto = linha[3].value
    pyperclip.copy(codigo_do__produto)
    pyautogui.click(134,516,duration=1) 
    pyautogui.hotkey('ctrl', 'v')
    
    peso_kg = linha[4].value
    pyperclip.copy(peso_kg)
    pyautogui.click(138,595,duration=1) 
    pyautogui.hotkey('ctrl', 'v')
    
    dimensao = linha[5].value
    pyperclip.copy(dimensao)
    pyautogui.click(139,673,duration=1) 
    pyautogui.hotkey('ctrl', 'v')
    
    pyautogui.click(152,724,duration=1)       # clicar no próximo
    sleep(3)
    
    preco = linha[6].value
    pyperclip.copy(preco)
    pyautogui.click(273,260,duration=1) 
    pyautogui.hotkey('ctrl', 'v')
    
    quantidade_em_estoque = linha[7].value
    pyperclip.copy(quantidade_em_estoque)
    pyautogui.click(229,340,duration=1) 
    pyautogui.hotkey('ctrl', 'v') 

    data_de_validade = linha[8].value
    pyperclip.copy(data_de_validade)
    pyautogui.click(196,414,duration=1) 
    pyautogui.hotkey('ctrl', 'v')
    
    cor = linha[9].value
    pyperclip.copy(cor)
    pyautogui.click(197,495,duration=1) 
    pyautogui.hotkey('ctrl', 'v')
    
    tamanho = linha[10].value
    pyautogui.click(189,573,duration=1)# Observaçaõ: ler inforomação de tamanho na planilha, pois tem 3 tamanhos pequeno, médio e grande
    if tamanho == "Pequeno":
        pyautogui.click(169,600,duration=1)   
    elif tamanho == "Médio":
        pyautogui.click(166,621,duration=1)  
    else: 
        pyautogui.click(172,637,duration=1)
          
    # pyperclip.copy(tamanho)
    # pyautogui.click(152,497,duration=1) 
    # pyautogui.hotkey('ctrl', 'v')
    
    material = linha[11].value
    pyperclip.copy(material)
    pyautogui.click(188,650,duration=1) 
    pyautogui.hotkey('ctrl', 'v')
    
    pyautogui.click(159,701,duration=1)       # clicar no próximo
    sleep(3)
    
    fabricante = linha[12].value
    pyperclip.copy(fabricante)
    pyautogui.click(166,277,duration=1) 
    pyautogui.hotkey('ctrl', 'v')

    pais_origem = linha[13].value
    pyperclip.copy(pais_origem)
    pyautogui.click(157,356,duration=1) 
    pyautogui.hotkey('ctrl', 'v')
    
    observacoes = linha[14].value
    pyperclip.copy(observacoes)
    pyautogui.click(163,456,duration=1) 
    pyautogui.hotkey('ctrl', 'v')
    
    codigo_de_barras = linha[15].value
    pyperclip.copy(codigo_de_barras)
    pyautogui.click(165,556,duration=1) 
    pyautogui.hotkey('ctrl', 'v')
    
    localizacao_no_armazem = linha[16].value
    pyperclip.copy(localizacao_no_armazem)
    pyautogui.click(158,631,duration=1) 
    pyautogui.hotkey('ctrl', 'v')
    
    pyautogui.click(162,685,duration=1)       # Concluido
    sleep(3)
    
    pyautogui.click(883,187,duration=1)       # Produto salvo no banco de dados!
    sleep(3)
    
    pyautogui.click(697,487,duration=1)       # Obrigado por registrar um novo produto. Se deseja adicionar outro, clique no botão abaixo.
    sleep(3)